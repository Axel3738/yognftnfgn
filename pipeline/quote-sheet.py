#!/usr/bin/env python3
"""Genererar ett Supplier Quote Request-sheet med exakt samma struktur som Axels mall.

    python3 pipeline/quote-sheet.py produkter.json ut.xlsx

produkter.json:
    [
      {"namn": "Makita 18V LED-arbetslampa",          # valfritt - mallen har tomt namnfält
       "temu_lank": "https://www.temu.com/se/...",     # obligatoriskt
       "bild": "bilder/lampa.png",                     # valfri lokal bildfil
       "butikslank": "",                               # Bäverbutiken-länk, oftast tom
       "leverantor_ref": ""}                           # Supplier ref / variant
    ]

Struktur (avläst ur mallen 2026-08-26): titel A1:AN1, instruktion A2:AN2,
tvaradig rubrik 3-4, darefter ett produktblock var 4:e rad (3 rader = kvantitet
1/2/3 + en tom mellanrad). Gula celler = leverantoren fyller i. Kraver openpyxl.
"""
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "Carlito"
# Fargerna ar avlasta ur mallen - andra dem inte utan att mallen andrats.
C_TITEL, C_INSTR = "FF17324D", "FFEAF1F5"
C_RUBRIK, C_UNDERRUBRIK = "FF244E66", "FFD8E0E6"
C_PRODUKT_A, C_PRODUKT_B = "FFD9EAF2", "FFE5EDF3"   # blocken vaxlar ton
C_LJUS_A, C_LJUS_B = "FFF4F7FA", "FFF7FBFD"
C_GUL, C_GRA, C_GRON = "FFFFF2CC", "FFE7EAED", "FFD9EAD3"

TITEL = "SUPPLIER QUOTE REQUEST — PRODUCT & SHIPPING COSTS"
INSTR = ("Please fill all yellow cells. No prices have been pre-filled. Enter product cost, "
         "shipping cost, delivery time and shipping method for quantities 1, 2 and 3 to each market.")

# (kolumn, rubrik) for det inledande blocket. None = ingen egen rubrik (del av merge).
HUVUD = [("A", "Product"), ("D", "QUOTE IMAGE / NOTES"), ("G", "Sale price on my store\n(leave blank)"),
         ("H", "MARKET"), ("I", "Product cost"), ("J", "Shipping cost"), ("K", "Total tax exclusive"),
         ("L", "Bäverbutiken link"), ("M", "TEMU LINK"), ("P", "Supplier ref / variant")]
HUVUD_MERGE = {"A": "A{r}:C{r2}", "D": "D{r}:F{r2}", "G": "G{r}:G{r2}", "H": "H{r}:H{r2}",
               "I": "I{r}:I{r2}", "J": "J{r}:J{r2}", "K": "K{r}:K{r2}", "L": "L{r}:L{r2}",
               "M": "M{r}:O{r2}", "P": "P{r}:P{r2}"}
# Exportmarknader: (rubrik, forsta kolumn). Sex kolumner var.
# (rubrik, forsta kolumn, rubrikfarg, qty-ton) - varje marknad har egna farger i mallen
MARKNADER = [("NORWAY", "Q", "FF527CA5", "FFDCEAF7"), ("FINLAND", "W", "FF6C6294", "FFC9DAF8"),
             ("DENMARK", "AC", "FFB17C2D", "FFFCE5CD"), ("UK", "AI", "FF9B5D6C", "FFD9D2E9")]
UNDERRUBRIKER = ["Qty", "Product cost", "Shipping cost", "Total ex. tax", "Delivery time", "Shipping method"]
# Exakt de bredder mallen sätter - ovriga kolumner lamnas pa standardbredd.
BREDDER = {"A": 18, "B": 4, "D": 12, "E": 7, "G": 14, "H": 11, "K": 13, "L": 20, "M": 13, "N": 8, "P": 27.4,
           "Q": 10, "T": 11, "U": 14,
           "W": 6, "X": 10, "Z": 11, "AA": 14,
           "AC": 6, "AD": 10, "AF": 11, "AG": 14,
           "AI": 6, "AJ": 10, "AL": 11, "AM": 14}
HOJD_PRODUKTRAD = 45     # mallen varierar (45/22.5/15.75); 45 valt sa bilden alltid far plats
HOJD_MELLANRAD = 6.75

KANT = Border(*[Side(style="thin", color="B7C4CE")] * 4)
MITT = Alignment(horizontal="center", vertical="center", wrap_text=True)
VANSTER = Alignment(horizontal="left", vertical="center", wrap_text=True)


def fyll(f):
    return PatternFill("solid", fgColor=f)


def kol(bokstav, steg=0):
    """Kolumnbokstav 'steg' kolumner till hoger om 'bokstav'."""
    from openpyxl.utils import column_index_from_string
    return get_column_letter(column_index_from_string(bokstav) + steg)


def bygg(produkter, ut, tomma=2):
    wb = Workbook()
    ws = wb.active
    ws.title = "Supplier Quote"
    sista = kol("AI", 5)  # AN

    ws.merge_cells(f"A1:{sista}1")
    ws["A1"] = TITEL
    ws["A1"].font = Font(name=FONT, size=16, bold=True, color="FFFFFFFF")
    ws["A1"].fill = fyll(C_TITEL)
    ws["A1"].alignment = VANSTER
    ws.row_dimensions[1].height = 27.75

    ws.merge_cells(f"A2:{sista}2")
    ws["A2"] = INSTR
    ws["A2"].font = Font(name=FONT, size=10)
    ws["A2"].fill = fyll(C_INSTR)
    ws["A2"].alignment = VANSTER
    ws.row_dimensions[2].height = 30

    for r in (3, 4):
        ws.row_dimensions[r].height = 25.5
    for c, rubrik in HUVUD:
        ws.merge_cells(HUVUD_MERGE[c].format(r=3, r2=4))
        cell = ws[f"{c}3"]
        cell.value = rubrik
        cell.font = Font(name=FONT, size=11, bold=True, color="FFFFFFFF")
        cell.fill = fyll(C_RUBRIK)
        cell.alignment = MITT
    for namn, start, farg, _q in MARKNADER:
        ws.merge_cells(f"{start}3:{kol(start, 5)}3")
        cell = ws[f"{start}3"]
        cell.value = namn
        cell.font = Font(name=FONT, size=11, bold=True, color="FFFFFFFF")
        cell.fill = fyll(farg)
        cell.alignment = MITT
        for i, under in enumerate(UNDERRUBRIKER):
            u = ws[f"{kol(start, i)}4"]
            u.value = under
            u.font = Font(name=FONT, size=9, bold=True)
            u.fill = fyll(C_UNDERRUBRIK)
            u.alignment = MITT
            u.border = KANT

    for b, bredd in BREDDER.items():
        ws.column_dimensions[b].width = bredd

    rad = 5
    for blocknr, p in enumerate(list(produkter) + [None] * tomma):
        c_produkt = C_PRODUKT_A if blocknr % 2 == 0 else C_PRODUKT_B
        c_ljus = C_LJUS_A if blocknr % 2 == 0 else C_LJUS_B
        if p is None:
            p = {"temu_lank": None}
        r2 = rad + 2
        for c in HUVUD_MERGE:
            if c in ("I", "J", "K"):   # en cell per kvantitetsrad - slas inte ihop
                continue
            ws.merge_cells(HUVUD_MERGE[c].format(r=rad, r2=r2))
        for _, start, _f, _q in MARKNADER:     # Delivery time och Shipping method galler hela produkten
            for j in (4, 5):
                ws.merge_cells(f"{kol(start, j)}{rad}:{kol(start, j)}{r2}")

        a = ws[f"A{rad}"]
        a.value = p.get("namn") or None
        a.fill = fyll(c_produkt)
        a.font = Font(name=FONT, size=10, bold=True)
        a.alignment = MITT
        for c, f in (("D", C_GUL), ("G", C_GRA), ("L", c_ljus), ("P", C_GUL)):
            cell = ws[f"{c}{rad}"]
            cell.fill = fyll(f)
            cell.font = Font(name=FONT, size=11)
            cell.alignment = MITT
        ws[f"L{rad}"] = p.get("butikslank") or None
        ws[f"P{rad}"] = p.get("leverantor_ref") or None

        h = ws[f"H{rad}"]
        h.value = "SWEDEN"
        h.fill = fyll(c_ljus)
        h.font = Font(name=FONT, size=11)
        h.alignment = MITT

        m = ws[f"M{rad}"]
        m.value = p.get("temu_lank")
        m.font = Font(name=FONT, size=11)
        m.alignment = VANSTER

        for i in range(3):
            r = rad + i
            ws.row_dimensions[r].height = HOJD_PRODUKTRAD
            for c, f in (("I", C_GUL), ("J", C_GUL), ("K", C_GRON)):
                cell = ws[f"{c}{r}"]
                cell.fill = fyll(f)
                cell.font = Font(name=FONT, size=11)
                cell.border = KANT
                cell.alignment = MITT
            for _, start, _f, qton in MARKNADER:
                q = ws[f"{start}{r}"]
                q.value = i + 1
                q.fill = fyll(qton)
                q.font = Font(name=FONT, size=11)
                q.alignment = MITT
                q.border = KANT
                for j, f in ((1, C_GUL), (2, C_GUL), (3, C_GRON), (4, C_GUL), (5, C_GUL)):
                    cell = ws[f"{kol(start, j)}{r}"]
                    cell.fill = fyll(f)
                    cell.font = Font(name=FONT, size=11)
                    cell.alignment = MITT
                    cell.border = KANT

        bild = p.get("bild")
        if bild and Path(bild).exists():
            img = XLImage(bild)
            skala = min(120 / img.width, 170 / img.height)
            img.width, img.height = int(img.width * skala), int(img.height * skala)
            ws.add_image(img, f"A{rad}")

        ws.row_dimensions[rad + 3].height = HOJD_MELLANRAD
        rad += 4

    # Mellanraderna mellan blocken ar vita i mallen
    vit = fyll("FFFFFFFF")
    from openpyxl.utils import column_index_from_string
    sista_i = column_index_from_string(sista)
    for r in range(8, rad, 4):
        for c in range(1, sista_i + 1):
            ws.cell(r, c).fill = vit
    # Mallen barsamma fyllning pa varje cell i en merge, inte bara ankaret
    for omrade in list(ws.merged_cells.ranges):
        ankare = ws.cell(omrade.min_row, omrade.min_col)
        if not (ankare.fill and ankare.fill.patternType):
            continue
        f = fyll(ankare.fill.start_color.rgb)
        for r in range(omrade.min_row, omrade.max_row + 1):
            for c in range(omrade.min_col, omrade.max_col + 1):
                ws.cell(r, c).fill = f

    ws.sheet_view.showGridLines = False
    wb.save(ut)
    return rad - 4


if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    produkter = json.loads(Path(sys.argv[1]).read_text(encoding="utf-8"))
    n = bygg(produkter, sys.argv[2])
    print(f"Skrev {sys.argv[2]} med {len(produkter)} produkter (sista raden {n}).")
